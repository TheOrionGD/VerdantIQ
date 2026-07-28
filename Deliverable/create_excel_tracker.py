import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

def build_verdantiq_tracker():
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # STYLES & PALETTE
    # ----------------------------------------------------
    font_family = "Segoe UI"
    
    # Header Styles
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    title_font = Font(name=font_family, size=16, bold=True, color="0F766E")
    subtitle_font = Font(name=font_family, size=11, italic=True, color="4B5563")
    section_font = Font(name=font_family, size=12, bold=True, color="111827")
    bold_font = Font(name=font_family, size=10, bold=True)
    regular_font = Font(name=font_family, size=10)
    
    # Fills
    primary_header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid") # Forest Teal
    dark_navy_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")      # Dark Slate
    sub_header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")     # Medium Slate
    card_bg_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")        # Soft Slate Tint
    
    # Status Fills & Fonts
    fill_done = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    font_done = Font(name=font_family, size=10, bold=True, color="065F46")
    
    fill_inflight = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    font_inflight = Font(name=font_family, size=10, bold=True, color="92400E")
    
    fill_notstarted = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    font_notstarted = Font(name=font_family, size=10, color="6B7280")
    
    fill_critical = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    font_critical = Font(name=font_family, size=10, bold=True, color="991B1B")

    fill_high = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
    font_high = Font(name=font_family, size=10, bold=True, color="9A3412")

    fill_medium = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    font_medium = Font(name=font_family, size=10, color="075985")
    
    fill_low = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    font_low = Font(name=font_family, size=10, color="475569")

    # Borders
    thin_border_side = Side(style='thin', color='CBD5E1')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom = Border(bottom=Side(style='medium', color='0F766E'))

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')

    # ====================================================
    # TAB 1: EXECUTIVE DASHBOARD & TEAM
    # ====================================================
    ws_dash = wb.active
    ws_dash.title = "Overview & Team Matrix"
    ws_dash.views.sheetView[0].showGridLines = True
    
    ws_dash.cell(row=2, column=2, value="🌱 VerdantIQ (EcoSphere) — System Tracking & Management Hub").font = title_font
    ws_dash.cell(row=3, column=2, value="Comprehensive Project Roadmap, Workflow Stage Checklist, System Bug Tracking & Component Registry").font = subtitle_font
    
    # Project Summary Card
    ws_dash.cell(row=5, column=2, value="PROJECT METADATA").font = section_font
    metadata = [
        ("Project Name:", "VerdantIQ (EcoSphere) Sustainability Intelligence Platform"),
        ("System Lead / Solution Architect:", "Grish Narayanan S"),
        ("Frontend & UI/UX Engineer:", "Gangash"),
        ("Software Developer & Tester:", "Girijesh"),
        ("Integration Support:", "Godfrey"),
        ("Tech Stack:", "Spring Boot 3.2, FastAPI, XGBoost, Google OR-Tools MILP, Groq AI, React 18, MongoDB, MinIO S3"),
        ("System Status:", "All 7 Core Phases Completed & Integration Tested ✅")
    ]
    
    curr_row = 6
    for label, val in metadata:
        c1 = ws_dash.cell(row=curr_row, column=2, value=label)
        c1.font = bold_font
        c1.fill = card_bg_fill
        c1.border = thin_border
        
        c2 = ws_dash.cell(row=curr_row, column=3, value=val)
        c2.font = regular_font
        c2.fill = card_bg_fill
        c2.border = thin_border
        curr_row += 1

    # Team Responsibility Matrix Table
    curr_row += 2
    ws_dash.cell(row=curr_row, column=2, value="👥 TEAM RESPONSIBILITY MATRIX").font = section_font
    curr_row += 1
    
    team_headers = ["Team Member", "Role", "Core Responsibilities", "Assigned Modules & Modules Handled"]
    for col_idx, h in enumerate(team_headers, start=2):
        cell = ws_dash.cell(row=curr_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = primary_header_fill
        cell.alignment = align_left
        cell.border = thin_border
    
    curr_row += 1
    team_data = [
        ("Grish Narayanan S", "Solution Architect & Backend Developer (Project Lead)", 
         "Design overall system architecture, develop backend services & REST APIs, manage database schemas (MongoDB), integrate ML/AI components, oversee technical decisions, and lead software development.",
         "Spring Boot Core Backend, MongoDB 2dsphere GIS, FastAPI XGBoost Pipeline, Google OR-Tools MILP Solver, Groq AI LLM Gateway"),
        
        ("Gangash", "UI/UX Engineer & Frontend Developer",
         "Design the UI/UX, develop frontend pages, implement responsive interfaces, 2D Spatial Canvas, dashboards, charts, form controls, mobile/desktop view modes, and ensure smooth user experience.",
         "React Spatial Canvas, Emerald Oasis Theme, Omnibar Command Center, Desktop/Mobile Views, 12 Workspace Nodes, Data Visualization Components"),
        
        ("Girijesh", "Software Developer & Tester",
         "Develop assigned application modules, assist in backend/frontend implementation, perform functional testing, integration testing, bug fixing, automated test suite maintenance, and quality assurance.",
         "JUnit 5 REST Test Suite, Pytest ML Test Suite, Node 36-Endpoint Integration Suite, Bug Remediation, Telemetry Monitoring"),
        
        ("Godfrey", "Integration Support",
         "Support frontend-backend integration, assist with AI/ML integration, help resolve cross-origin & storage issues, contribute to documentation, version control, deployment scripts, and technical support.",
         "MinIO S3 Object Storage, CORS Filters, PowerShell Multi-Service Launcher (run-phase11-system.ps1), PDF Statement Generator, Deployment Specs")
    ]
    
    for row_data in team_data:
        for col_idx, val in enumerate(row_data, start=2):
            cell = ws_dash.cell(row=curr_row, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx == 2:
                cell.font = bold_font
        ws_dash.row_dimensions[curr_row].height = 42
        curr_row += 1

    # Key Performance Indicators Summary
    curr_row += 2
    ws_dash.cell(row=curr_row, column=2, value="📈 SYSTEM HEALTH & EXECUTION METRICS").font = section_font
    curr_row += 1
    
    metrics_headers = ["Metric Category", "Target Scope", "Completed / Verified", "Health Status"]
    for col_idx, h in enumerate(metrics_headers, start=2):
        cell = ws_dash.cell(row=curr_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = sub_header_fill
        cell.alignment = align_left
        cell.border = thin_border
        
    curr_row += 1
    metrics_data = [
        ("Core Development Phases", "7 Phases", "7 Completed (100%)", "DONE ✅"),
        ("System REST & ML Endpoints", "36 Endpoints", "36 Tested & Verified (100%)", "DONE ✅"),
        ("Spatial UI Workspace Nodes", "12 Nodes", "12 Fully Interactive Nodes", "DONE ✅"),
        ("Automated Test Suites", "3 Suites (JUnit, Pytest, Integration)", "All Passing Cleanly", "DONE ✅"),
        ("Bug Resolution & Refactoring", "7 System Bugs Logged", "7 Resolved / Closed (100%)", "DONE ✅"),
        ("Production Script Launcher", "PowerShell Launch Script", "Configured & Tested", "DONE ✅")
    ]
    for m_cat, m_tgt, m_cmp, m_stat in metrics_data:
        c2 = ws_dash.cell(row=curr_row, column=2, value=m_cat)
        c3 = ws_dash.cell(row=curr_row, column=3, value=m_tgt)
        c4 = ws_dash.cell(row=curr_row, column=4, value=m_cmp)
        c5 = ws_dash.cell(row=curr_row, column=5, value=m_stat)
        
        for c in [c2, c3, c4, c5]:
            c.font = regular_font
            c.border = thin_border
        c2.font = bold_font
        c5.font = font_done
        c5.fill = fill_done
        c5.alignment = align_center
        curr_row += 1


    # ====================================================
    # TAB 2: GANTT CHART & ROADMAP
    # ====================================================
    ws_gantt = wb.create_sheet(title="🗓️ Gantt Chart & Roadmap")
    ws_gantt.views.sheetView[0].showGridLines = True
    
    ws_gantt.cell(row=2, column=2, value="🗓️ VerdantIQ System Development Roadmap & Gantt Schedule").font = title_font
    ws_gantt.cell(row=3, column=2, value="Phase-by-phase timeline, task ownership, duration in days, completion percentage, and 12-week schedule grid").font = subtitle_font
    
    gantt_headers = [
        "Task ID", "Phase / Task Description", "Assigned Lead", "Start Date", "Due Date", "Duration (Days)", "% Complete",
        "W1", "W2", "W3", "W4", "W5", "W6", "W7", "W8", "W9", "W10", "W11", "W12"
    ]
    
    header_row_gantt = 5
    for col_idx, h in enumerate(gantt_headers, start=2):
        cell = ws_gantt.cell(row=header_row_gantt, column=col_idx, value=h)
        cell.font = header_font
        cell.alignment = align_center if col_idx >= 9 else align_left
        cell.border = thin_border
        if col_idx >= 9:
            cell.fill = dark_navy_fill
        else:
            cell.fill = primary_header_fill

    gantt_tasks = [
        # (ID, Title, Lead, Start, Due, Duration, Pct, ActiveWeeks)
        ("1.0", "PHASE 1: Client UI/UX & Spatial Engine", "Gangash", "2026-06-01", "2026-06-14", 14, 1.0, [1, 2]),
        ("1.1", "  - Emerald Oasis Design System & Theme", "Gangash", "2026-06-01", "2026-06-04", 4, 1.0, [1]),
        ("1.2", "  - React 2D Infinite Spatial Canvas Shell", "Gangash", "2026-06-05", "2026-06-09", 5, 1.0, [1, 2]),
        ("1.3", "  - Groq AI Omnibar & 12 Workspace Nodes", "Gangash", "2026-06-10", "2026-06-14", 5, 1.0, [2]),
        
        ("2.0", "PHASE 2: Backend REST Services (Spring Boot 3.2)", "Grish Narayanan S", "2026-06-15", "2026-06-28", 14, 1.0, [3, 4]),
        ("2.1", "  - JWT Auth Security & User Profile Service", "Grish Narayanan S", "2026-06-15", "2026-06-18", 4, 1.0, [3]),
        ("2.2", "  - Activity Tracker & MongoDB CRUD Endpoints", "Grish Narayanan S", "2026-06-19", "2026-06-23", 5, 1.0, [3, 4]),
        ("2.3", "  - Apache PDFBox Monthly Statement Generator", "Grish Narayanan S", "2026-06-24", "2026-06-28", 5, 1.0, [4]),
        
        ("3.0", "PHASE 3: ML Forecasting & OR-Tools Optimization", "Grish Narayanan S", "2026-06-29", "2026-07-12", 14, 1.0, [5, 6]),
        ("3.1", "  - FastAPI Pipeline & XGBoost 30-Day Forecast", "Grish Narayanan S", "2026-06-29", "2026-07-03", 5, 1.0, [5]),
        ("3.2", "  - IsolationForest Spike Anomaly Engine", "Girijesh", "2026-07-04", "2026-07-07", 4, 1.0, [5, 6]),
        ("3.3", "  - Google OR-Tools MILP Constraint Solver", "Grish Narayanan S", "2026-07-08", "2026-07-12", 5, 1.0, [6]),
        
        ("4.0", "PHASE 4: Geospatial GIS & Storage Tier", "Godfrey", "2026-07-13", "2026-07-19", 7, 1.0, [7]),
        ("4.1", "  - MongoDB 2dsphere GeoJSON Geofencing", "Grish Narayanan S", "2026-07-13", "2026-07-16", 4, 1.0, [7]),
        ("4.2", "  - MinIO S3 Proof Upload & Tesseract OCR Integration", "Godfrey", "2026-07-17", "2026-07-19", 3, 1.0, [7]),

        ("5.0", "PHASE 5: Generative AI & Groq LLM Gateway", "Grish Narayanan S", "2026-07-20", "2026-07-23", 4, 1.0, [8]),
        ("5.1", "  - Grounded System Prompt & Chat Endpoint", "Grish Narayanan S", "2026-07-20", "2026-07-21", 2, 1.0, [8]),
        ("5.2", "  - Local Grounded Fallback Engine & Guardrails", "Godfrey", "2026-07-22", "2026-07-23", 2, 1.0, [8]),

        ("6.0", "PHASE 6: Notifications SSE & System Telemetry", "Girijesh", "2026-07-24", "2026-07-26", 3, 1.0, [9]),
        ("6.1", "  - Real-time Server-Sent Events (SSE) Stream", "Girijesh", "2026-07-24", "2026-07-25", 2, 1.0, [9]),
        ("6.2", "  - System Health & Telemetry Metrics Endpoint", "Girijesh", "2026-07-25", "2026-07-26", 2, 1.0, [9]),

        ("7.0", "PHASE 7: Automated Verification & Production Release", "Girijesh", "2026-07-27", "2026-07-28", 2, 1.0, [10, 11, 12]),
        ("7.1", "  - JUnit 5 + Pytest + 36-Endpoint Integration Suite", "Girijesh", "2026-07-27", "2026-07-27", 1, 1.0, [10]),
        ("7.2", "  - PowerShell Launch Automation (run-phase11-system.ps1)", "Godfrey", "2026-07-28", "2026-07-28", 1, 1.0, [11, 12])
    ]

    curr_row_gantt = 6
    gantt_bar_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid") # Mint emerald bar
    parent_bar_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid") # Darker teal parent bar
    
    for item in gantt_tasks:
        tid, desc, lead, sdate, edate, dur, pct, w_list = item
        is_parent = tid.endswith(".0")
        
        c_id = ws_gantt.cell(row=curr_row_gantt, column=2, value=tid)
        c_desc = ws_gantt.cell(row=curr_row_gantt, column=3, value=desc)
        c_lead = ws_gantt.cell(row=curr_row_gantt, column=4, value=lead)
        c_sdate = ws_gantt.cell(row=curr_row_gantt, column=5, value=sdate)
        c_edate = ws_gantt.cell(row=curr_row_gantt, column=6, value=edate)
        c_dur = ws_gantt.cell(row=curr_row_gantt, column=7, value=dur)
        c_pct = ws_gantt.cell(row=curr_row_gantt, column=8, value=pct)
        c_pct.number_format = '0%'
        
        for c in [c_id, c_desc, c_lead, c_sdate, c_edate, c_dur, c_pct]:
            c.border = thin_border
            if is_parent:
                c.font = bold_font
                c.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            else:
                c.font = regular_font
                
        # Fill Gantt bar columns (Columns 9 to 20 for W1..W12)
        for w_idx in range(1, 13):
            bar_cell = ws_gantt.cell(row=curr_row_gantt, column=8 + w_idx)
            bar_cell.border = thin_border
            if w_idx in w_list:
                bar_cell.fill = parent_bar_fill if is_parent else gantt_bar_fill
            elif is_parent:
                bar_cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                
        curr_row_gantt += 1


    # ====================================================
    # TAB 3: WORKFLOW & STAGE CHECKLIST
    # ====================================================
    ws_wf = wb.create_sheet(title="📋 Stage & Workflow Checklist")
    ws_wf.views.sheetView[0].showGridLines = True

    ws_wf.cell(row=2, column=2, value="📋 VerdantIQ Stage & Workflow Execution Checklist").font = title_font
    ws_wf.cell(row=3, column=2, value="Granular task breakdown across system stages: Design, Dev/Build, AI/ML, Testing & Release").font = subtitle_font
    
    wf_headers = ["Stage", "Sub-Stage / Area", "Step / Action Item", "Status", "Owner", "Priority", "Comment / Verification Note"]
    
    header_row_wf = 5
    for col_idx, h in enumerate(wf_headers, start=2):
        cell = ws_wf.cell(row=header_row_wf, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = primary_header_fill
        cell.alignment = align_left
        cell.border = thin_border

    checklist_data = [
        # PRE-RELEASE & SYSTEM DESIGN
        ("PRE-RELEASE", "System Design", "Define project goals, scope, and technical roadmap", "Done", "Grish Narayanan S", "High", "Roadmap approved in Deliverable/Readme.md"),
        ("PRE-RELEASE", "System Design", "Establish MongoDB GeoJSON schemas & ER diagrams", "Done", "Grish Narayanan S", "High", "Documented in SCHEMA.md with 2dsphere indexes"),
        ("PRE-RELEASE", "System Design", "Formulate OR-Tools MILP cost vs CO2 math model", "Done", "Grish Narayanan S", "Critical", "Formulated with pywraplp SCIP solver"),
        ("PRE-RELEASE", "UI/UX Design", "Design Emerald Oasis design system & palette", "Done", "Gangash", "High", "#059669 Mint, #0F766E Teal, #F3F7F5 Sage"),

        # FRONTEND DEVELOPMENT
        ("DEVELOPMENT", "Frontend UI", "Develop React 18 2D Spatial Canvas & nodes", "Done", "Gangash", "Critical", "Pan, zoom, & SVG glowing data streams active"),
        ("DEVELOPMENT", "Frontend UI", "Implement Groq AI Floating Omnibar Command Center", "Done", "Gangash", "High", "Natural language intent parser integrated"),
        ("DEVELOPMENT", "Frontend UI", "Build 3-Mode View Switcher (Canvas, Desktop, Mobile)", "Done", "Gangash", "High", "Seamless responsive view toggle verified"),
        ("DEVELOPMENT", "Frontend UI", "Build 12 Interactive Workspace Nodes & Cards", "Done", "Gangash", "High", "Dashboard, Tracker, Twin, Optimizer, Admin, etc."),

        # BACKEND DEVELOPMENT
        ("DEVELOPMENT", "Backend API", "Configure Spring Boot 3.2 + Spring Security JWT", "Done", "Grish Narayanan S", "Critical", "Role-based access (USER, INSTITUTION_ADMIN, SYSTEM_ADMIN)"),
        ("DEVELOPMENT", "Backend API", "Implement Activity Tracker Service & GeoJSON Logging", "Done", "Grish Narayanan S", "High", "/api/v1/tracker endpoints with spatial queries"),
        ("DEVELOPMENT", "Backend API", "Build Institutional Leaderboard & Aggregation Pipelines", "Done", "Grish Narayanan S", "High", "MongoDB $group and $sum aggregations active"),
        ("DEVELOPMENT", "Backend API", "Implement Apache PDFBox Statement Service", "Done", "Grish Narayanan S", "Medium", "/api/v1/reports PDF downloads operational"),

        # ML & AI ENGINES
        ("DEVELOPMENT", "ML & AI Engine", "Develop FastAPI XGBoost 30-Day Resource Forecaster", "Done", "Grish Narayanan S", "Critical", "/api/v1/ml/forecasts running on port 8000"),
        ("DEVELOPMENT", "ML & AI Engine", "Develop Scikit-learn IsolationForest Anomaly Engine", "Done", "Girijesh", "High", "/api/v1/ml/explain spike detection active"),
        ("DEVELOPMENT", "ML & AI Engine", "Implement Google OR-Tools MILP Solver Endpoint", "Done", "Grish Narayanan S", "Critical", "/api/v1/optimizer/solve handling budgets & targets"),
        ("DEVELOPMENT", "ML & AI Engine", "Integrate Groq API Llama-3.3 LLM Assistant Endpoint", "Done", "Grish Narayanan S", "Critical", "/api/v1/assistant/chat with grounded context"),
        ("DEVELOPMENT", "ML & AI Engine", "Build Local Grounded Fallback Engine for offline LLM", "Done", "Godfrey", "High", "Ensures zero service interruption on rate-limit"),

        # GIS & STORAGE TIER
        ("DEVELOPMENT", "GIS & Storage", "Configure MinIO / AWS S3 presigned proof storage", "Done", "Godfrey", "High", "Bill upload & evidence storage connected"),
        ("DEVELOPMENT", "GIS & Storage", "Integrate Tesseract OCR for bill text parsing", "Done", "Godfrey", "Medium", "Automated kilowatt-hour & cost extraction"),
        ("DEVELOPMENT", "GIS & Storage", "Implement MongoDB $geoWithin Geofence Verification", "Done", "Grish Narayanan S", "High", "Campus radius verification active"),

        # TESTING & QA
        ("QUALITY ASSURANCE", "Testing & QA", "Write JUnit 5 Spring Security & REST Controller tests", "Done", "Girijesh", "High", "Auth, User, Tracker, and Report tests passing"),
        ("QUALITY ASSURANCE", "Testing & QA", "Write Pytest XGBoost & OR-Tools Solver tests", "Done", "Girijesh", "High", "MILP boundary & forecasting tests passing"),
        ("QUALITY ASSURANCE", "Testing & QA", "Build Node.js 36-Endpoint Integration Test Suite", "Done", "Girijesh", "Critical", "Comprehensive end-to-end integration verified"),

        # DEPLOYMENT & DEVOPS
        ("RELEASE / DEVOPS", "Deployment", "Create PowerShell Multi-Service System Launcher", "Done", "Godfrey", "High", "run-phase11-system.ps1 launches all servers"),
        ("RELEASE / DEVOPS", "Deployment", "Prepare Production Deployment Manifest & Configs", "Done", "Godfrey", "High", "DEPLOYMENT.md ready for Railway, Render, & Vercel")
    ]

    curr_row_wf = 6
    for item in checklist_data:
        stg, sub, step, stat, owner, prio, comm = item
        
        c_stg = ws_wf.cell(row=curr_row_wf, column=2, value=stg)
        c_sub = ws_wf.cell(row=curr_row_wf, column=3, value=sub)
        c_step = ws_wf.cell(row=curr_row_wf, column=4, value=step)
        c_stat = ws_wf.cell(row=curr_row_wf, column=5, value=stat)
        c_own = ws_wf.cell(row=curr_row_wf, column=6, value=owner)
        c_prio = ws_wf.cell(row=curr_row_wf, column=7, value=prio)
        c_comm = ws_wf.cell(row=curr_row_wf, column=8, value=comm)
        
        for c in [c_stg, c_sub, c_step, c_stat, c_own, c_prio, c_comm]:
            c.border = thin_border
            c.font = regular_font
            
        c_stg.font = bold_font
        c_stat.alignment = align_center
        c_prio.alignment = align_center
        
        # Status styling
        if stat == "Done":
            c_stat.fill = fill_done
            c_stat.font = font_done
        elif stat == "In Flight":
            c_stat.fill = fill_inflight
            c_stat.font = font_inflight
        else:
            c_stat.fill = fill_notstarted
            c_stat.font = font_notstarted
            
        # Priority styling
        if prio == "Critical":
            c_prio.fill = fill_critical
            c_prio.font = font_critical
        elif prio == "High":
            c_prio.fill = fill_high
            c_prio.font = font_high
        elif prio == "Medium":
            c_prio.fill = fill_medium
            c_prio.font = font_medium
        else:
            c_prio.fill = fill_low
            c_prio.font = font_low

        curr_row_wf += 1


    # ====================================================
    # TAB 4: SYSTEM BUG & CHANGE LOG (GIT COMMITS & ISSUE TRACKING)
    # ====================================================
    ws_bug = wb.create_sheet(title="🐛 System Bug & Change Log")
    ws_bug.views.sheetView[0].showGridLines = True

    ws_bug.cell(row=2, column=2, value="🐛 VerdantIQ System Bug, Defect & Architectural Change Log").font = title_font
    ws_bug.cell(row=3, column=2, value="Git commit-level issue tracking, root cause analysis, developer assignment, and resolution history").font = subtitle_font
    
    bug_headers = [
        "Bug / Ticket ID", "Type", "Description & Root Cause", "Date Identified", "Status", "Priority", 
        "Assigned Developer", "Git Commit / Ref", "Impacted Module", "Action Taken & Fix", "Escalation", "Date Resolved"
    ]
    
    header_row_bug = 5
    for col_idx, h in enumerate(bug_headers, start=2):
        cell = ws_bug.cell(row=header_row_bug, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = dark_navy_fill
        cell.alignment = align_left
        cell.border = thin_border

    bug_data = [
        ("BUG-001", "Defect", "OR-Tools MILP Solver failed with INFEASIBLE status when budget constraint was set to 0", "2026-07-10", "Closed", "Critical",
         "Grish Narayanan S", "feat(optimizer): commit #a7f3d2", "FastAPI / OR-Tools Engine",
         "Added feasibility pre-check and explicit fallback response returning clean INFEASIBLE explanation payload.", "N", "2026-07-11"),

        ("BUG-002", "Defect", "Expired or malformed JWT token triggered 500 Internal Server Error instead of 401 Unauthorized", "2026-06-17", "Closed", "Critical",
         "Grish Narayanan S", "fix(auth): commit #c891e4", "Spring Boot Auth Filter",
         "Configured JwtAuthenticationEntryPoint in Spring Security filter chain to catch ExpiredJwtException.", "N", "2026-06-18"),

        ("BUG-003", "UI Defect", "SVG Canvas pipeline connectors misaligned on high-DPI / 4K display viewports", "2026-06-08", "Closed", "High",
         "Gangash", "fix(canvas): commit #b451d0", "React Spatial Canvas",
         "Refactored SVG path coordinate calculations using dynamic bounding box getters (getBoundingClientRect).", "N", "2026-06-09"),

        ("BUG-004", "Integration", "MinIO S3 presigned URL generation threw bucket not found error during initial proof upload", "2026-07-18", "Closed", "High",
         "Godfrey", "fix(storage): commit #e129f8", "MinIO S3 Service",
         "Added automatic bucket initialization check (bucketExists / makeBucket) at Spring Boot startup.", "N", "2026-07-18"),

        ("BUG-005", "Refactor", "Groq AI LLM external API timeout when user query was submitted during peak network latency", "2026-07-21", "Closed", "Critical",
         "Grish Narayanan S", "feat(ai): commit #d3389a", "Groq AI LLM Gateway",
         "Built Local Grounded Fallback Engine returning heuristic recommendations when Groq API exceeds 3000ms response window.", "N", "2026-07-22"),

        ("BUG-006", "Database", "Spatial query $geoWithin failed due to missing 2dsphere index on ActivityLog.location", "2026-07-14", "Closed", "High",
         "Grish Narayanan S", "fix(db): commit #f9801b", "MongoDB Database Tier",
         "Created @GeoSpatialIndexed(type = GeoSpatialIndexType.GEO_2DSPHERE) annotation on ActivityLog entity.", "N", "2026-07-14"),

        ("BUG-007", "Defect", "Server-Sent Events (SSE) stream closed automatically after 30 seconds idle state in browser", "2026-07-25", "Closed", "Medium",
         "Girijesh", "fix(notifications): commit #77c2e3", "Notification SSE Service",
         "Implemented scheduled heartbeat comment ping (:ping) every 15 seconds over SseEmitter channel.", "N", "2026-07-26")
    ]

    curr_row_bug = 6
    for item in bug_data:
        bid, btype, bdesc, bdate, bstat, bprio, bown, bgit, bmod, bfix, besc, bres = item
        
        c_id = ws_bug.cell(row=curr_row_bug, column=2, value=bid)
        c_type = ws_bug.cell(row=curr_row_bug, column=3, value=btype)
        c_desc = ws_bug.cell(row=curr_row_bug, column=4, value=bdesc)
        c_date = ws_bug.cell(row=curr_row_bug, column=5, value=bdate)
        c_stat = ws_bug.cell(row=curr_row_bug, column=6, value=bstat)
        c_prio = ws_bug.cell(row=curr_row_bug, column=7, value=bprio)
        c_own = ws_bug.cell(row=curr_row_bug, column=8, value=bown)
        c_git = ws_bug.cell(row=curr_row_bug, column=9, value=bgit)
        c_mod = ws_bug.cell(row=curr_row_bug, column=10, value=bmod)
        c_fix = ws_bug.cell(row=curr_row_bug, column=11, value=bfix)
        c_esc = ws_bug.cell(row=curr_row_bug, column=12, value=besc)
        c_res = ws_bug.cell(row=curr_row_bug, column=13, value=bres)

        for c in [c_id, c_type, c_desc, c_date, c_stat, c_prio, c_own, c_git, c_mod, c_fix, c_esc, c_res]:
            c.border = thin_border
            c.font = regular_font
            
        c_id.font = bold_font
        c_git.font = Font(name="Consolas", size=9.5, color="1E40AF")
        c_stat.alignment = align_center
        c_prio.alignment = align_center
        c_esc.alignment = align_center
        
        # Status
        c_stat.fill = fill_done
        c_stat.font = font_done
            
        # Priority
        if bprio == "Critical":
            c_prio.fill = fill_critical
            c_prio.font = font_critical
        elif bprio == "High":
            c_prio.fill = fill_high
            c_prio.font = font_high
        elif bprio == "Medium":
            c_prio.fill = fill_medium
            c_prio.font = font_medium
            
        ws_bug.row_dimensions[curr_row_bug].height = 32
        curr_row_bug += 1


    # ====================================================
    # TAB 5: SYSTEM MODULES & FEATURE REGISTRY
    # ====================================================
    ws_mod = wb.create_sheet(title="🗂️ System Modules & Registry")
    ws_mod.views.sheetView[0].showGridLines = True

    ws_mod.cell(row=2, column=2, value="🗂️ VerdantIQ Application Modules & Subsystem Registry").font = title_font
    ws_mod.cell(row=3, column=2, value="Overview of system microservices, frontend nodes, database collections, and technical leads").font = subtitle_font

    mod_headers = [
        "Module ID", "Module Name", "Layer / Service Tier", "Lead Owner", "Tech Stack & Libraries", 
        "Deployment / Route", "Stage", "Priority", "Test Coverage", "Key Capabilities"
    ]

    header_row_mod = 5
    for col_idx, h in enumerate(mod_headers, start=2):
        cell = ws_mod.cell(row=header_row_mod, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = primary_header_fill
        cell.alignment = align_left
        cell.border = thin_border

    modules_data = [
        ("MOD-001", "React Spatial Canvas Shell", "Frontend UI", "Gangash", "React 18, SVG Streams, Tailwind", "http://localhost:5173", "Production Ready", "Critical", "100%", "2D infinite canvas, drag-pan, zoom controls, glowing SVG streams"),
        ("MOD-002", "Groq AI Omnibar", "Frontend UI", "Gangash", "React Context, Web Speech API", "Global Floating Bar", "Production Ready", "High", "100%", "Natural language intent recognition, speech input, node navigation"),
        ("MOD-003", "Spring Boot Core REST Backend", "Backend Core", "Grish Narayanan S", "Java 21, Spring Boot 3.2, Security", "http://localhost:8080", "Production Ready", "Critical", "100%", "JWT authentication, user profiles, activity CRUD, PDF generation"),
        ("MOD-004", "FastAPI ML & Solver Engine", "ML Service", "Grish Narayanan S", "Python 3.11, FastAPI, Pydantic v2", "http://localhost:8000", "Production Ready", "Critical", "100%", "XGBoost forecasts, IsolationForest anomaly, OR-Tools solver"),
        ("MOD-005", "Google OR-Tools MILP Solver", "Optimization", "Grish Narayanan S", "Google OR-Tools (pywraplp)", "POST /api/v1/optimizer/solve", "Production Ready", "Critical", "100%", "Constrained linear programming for budget vs CO2 offset optimization"),
        ("MOD-006", "Groq AI LLM Gateway Service", "Generative AI", "Grish Narayanan S", "Groq API (Llama-3.3-70b)", "POST /api/v1/assistant/chat", "Production Ready", "Critical", "100%", "Grounded conversational decision intelligence with structured JSON output"),
        ("MOD-007", "MongoDB Geospatial Store", "Database Tier", "Grish Narayanan S", "MongoDB 7.0, 2dsphere index", "mongodb://localhost:27017", "Production Ready", "Critical", "100%", "GeoJSON point & polygon storage with $geoNear geofencing"),
        ("MOD-008", "MinIO S3 & OCR Proof Storage", "Object Storage", "Godfrey", "MinIO Java SDK, Tesseract OCR", "http://localhost:9000", "Production Ready", "High", "100%", "Presigned bill uploads, proof geotagging, automated text extraction"),
        ("MOD-009", "Real-Time Notifications SSE", "Notifications", "Girijesh", "Spring SseEmitter, Event Bus", "GET /api/v1/notifications/stream", "Production Ready", "Medium", "100%", "In-app notifications stream, heartbeat keep-alive, mark-as-read API"),
        ("MOD-010", "System Telemetry & Monitoring", "Monitoring", "Girijesh", "Spring Actuator, Custom Metrics", "GET /api/v1/admin/telemetry", "Production Ready", "Medium", "100%", "JVM memory, MongoDB ping latency, request throughput monitoring"),
        ("MOD-011", "PowerShell Launch Automation", "DevOps", "Godfrey", "PowerShell 7.x", "run-phase11-system.ps1", "Production Ready", "High", "100%", "Orchestrates Spring Boot, FastAPI, React UI, and Integration Tests"),
        ("MOD-012", "Node.js Integration Test Suite", "Testing & QA", "Girijesh", "Node.js, Axios, Assert", "test-system-endpoints.js", "Production Ready", "Critical", "100%", "36-endpoint automated verification covering all system endpoints")
    ]

    curr_row_mod = 6
    for item in modules_data:
        mid, mname, mlayer, mlead, mstack, mroute, mstage, mprio, mcov, mcap = item
        
        c_id = ws_mod.cell(row=curr_row_mod, column=2, value=mid)
        c_name = ws_mod.cell(row=curr_row_mod, column=3, value=mname)
        c_layer = ws_mod.cell(row=curr_row_mod, column=4, value=mlayer)
        c_lead = ws_mod.cell(row=curr_row_mod, column=5, value=mlead)
        c_stack = ws_mod.cell(row=curr_row_mod, column=6, value=mstack)
        c_route = ws_mod.cell(row=curr_row_mod, column=7, value=mroute)
        c_stage = ws_mod.cell(row=curr_row_mod, column=8, value=mstage)
        c_prio = ws_mod.cell(row=curr_row_mod, column=9, value=mprio)
        c_cov = ws_mod.cell(row=curr_row_mod, column=10, value=mcov)
        c_cap = ws_mod.cell(row=curr_row_mod, column=11, value=mcap)

        for c in [c_id, c_name, c_layer, c_lead, c_stack, c_route, c_stage, c_prio, c_cov, c_cap]:
            c.border = thin_border
            c.font = regular_font
            
        c_id.font = bold_font
        c_stage.alignment = align_center
        c_prio.alignment = align_center
        c_cov.alignment = align_center
        
        c_stage.fill = fill_done
        c_stage.font = font_done
        
        if mprio == "Critical":
            c_prio.fill = fill_critical
            c_prio.font = font_critical
        elif mprio == "High":
            c_prio.fill = fill_high
            c_prio.font = font_high
        else:
            c_prio.fill = fill_medium
            c_prio.font = font_medium
            
        curr_row_mod += 1


    # ----------------------------------------------------
    # AUTO-FIT COLUMN WIDTHS & PADDING FOR ALL SHEETS
    # ----------------------------------------------------
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            # Skip adjusting gantt week columns (Cols I to T) to keep them uniform
            if ws.title == "🗓️ Gantt Chart & Roadmap" and col[0].column >= 9:
                ws.column_dimensions[col_letter].width = 4.5
                continue
                
            for cell in col:
                val_str = str(cell.value or '')
                if cell.row in [2, 3]: # Skip titles
                    continue
                # Split line lengths for wrapped text
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            
            if max_len > 0:
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 65)

    # Specific column dimension overrides for perfect aesthetic alignment
    ws_dash.column_dimensions['A'].width = 3
    ws_dash.column_dimensions['B'].width = 35
    ws_dash.column_dimensions['C'].width = 50
    ws_dash.column_dimensions['D'].width = 45
    ws_dash.column_dimensions['E'].width = 25

    ws_gantt.column_dimensions['A'].width = 3
    ws_gantt.column_dimensions['B'].width = 10
    ws_gantt.column_dimensions['C'].width = 45
    ws_gantt.column_dimensions['D'].width = 22
    ws_gantt.column_dimensions['E'].width = 14
    ws_gantt.column_dimensions['F'].width = 14
    ws_gantt.column_dimensions['G'].width = 16
    ws_gantt.column_dimensions['H'].width = 14

    ws_wf.column_dimensions['A'].width = 3
    ws_wf.column_dimensions['B'].width = 22
    ws_wf.column_dimensions['C'].width = 25
    ws_wf.column_dimensions['D'].width = 48
    ws_wf.column_dimensions['E'].width = 14
    ws_wf.column_dimensions['F'].width = 22
    ws_wf.column_dimensions['G'].width = 14
    ws_wf.column_dimensions['H'].width = 45

    ws_bug.column_dimensions['A'].width = 3
    ws_bug.column_dimensions['B'].width = 16
    ws_bug.column_dimensions['C'].width = 14
    ws_bug.column_dimensions['D'].width = 48
    ws_bug.column_dimensions['E'].width = 14
    ws_bug.column_dimensions['F'].width = 12
    ws_bug.column_dimensions['G'].width = 14
    ws_bug.column_dimensions['H'].width = 20
    ws_bug.column_dimensions['I'].width = 25
    ws_bug.column_dimensions['J'].width = 25
    ws_bug.column_dimensions['K'].width = 45
    ws_bug.column_dimensions['L'].width = 12
    ws_bug.column_dimensions['M'].width = 14

    output_path = "o:\\PROJECTS\\VerdantIQ\\Deliverable\\VerdantIQ_System_Tracker_and_Roadmap.xlsx"
    wb.save(output_path)
    print(f"Successfully generated Excel file at: {output_path}")

if __name__ == "__main__":
    build_verdantiq_tracker()
